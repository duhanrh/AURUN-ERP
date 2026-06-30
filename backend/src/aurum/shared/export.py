"""Exportación tabular reutilizable a CSV, Excel (.xlsx) y PDF.

Genérico y desacoplado de cualquier módulo: opera sobre un ``ExportDoc`` (cabecera de
marca + columnas + filas + resumen). Lo usa Reportes hoy, pero cualquier módulo puede
construir un ``ExportDoc`` y reutilizar estos serializadores y ``export_response``.

PDF con fpdf2 usa fuentes núcleo (latin-1); ``_pdf_safe`` translitera/limpia los pocos
símbolos fuera de ese rango (p. ej. ``Δ``) para no requerir empacar una fuente TTF.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import datetime
from typing import Literal

from fastapi import Response

ExportFormat = Literal["csv", "xlsx", "pdf"]
EXPORT_FORMATS: tuple[ExportFormat, ...] = ("csv", "xlsx", "pdf")

_MEDIA = {
    "csv": "text/csv; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}


@dataclass(frozen=True, slots=True)
class ExportSummary:
    label: str
    value: str


@dataclass(frozen=True, slots=True)
class ExportDoc:
    """Documento tabular a exportar, con cabecera de marca."""

    brand_name: str
    title: str
    document_number: str
    generated_at: datetime
    columns: list[str]
    rows: list[list[str]]
    summary: list[ExportSummary] = field(default_factory=list)


# ── CSV ──────────────────────────────────────────────────────────────────────
def to_csv(doc: ExportDoc) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([doc.brand_name])
    writer.writerow([doc.title])
    writer.writerow(["Documento", doc.document_number])
    writer.writerow(["Generado", doc.generated_at.isoformat(timespec="seconds")])
    writer.writerow([])
    writer.writerow(doc.columns)
    writer.writerows(doc.rows)
    if doc.summary:
        writer.writerow([])
        writer.writerow(["Resumen"])
        for item in doc.summary:
            writer.writerow([item.label, item.value])
    return buffer.getvalue().encode("utf-8")


# ── Excel ────────────────────────────────────────────────────────────────────
def to_xlsx(doc: ExportDoc) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Reporte"

    gold = Font(bold=True, color="7A6228")
    header_fill = PatternFill("solid", fgColor="1A1A24")
    header_font = Font(bold=True, color="FFFFFF")

    ws["A1"] = doc.brand_name
    ws["A1"].font = Font(bold=True, size=14, color="7A6228")
    ws["A2"] = doc.title
    ws["A2"].font = gold
    ws["A3"] = f"Documento: {doc.document_number}"
    ws["A4"] = f"Generado: {doc.generated_at.isoformat(timespec='seconds')}"

    start = 6
    for col_idx, name in enumerate(doc.columns, start=1):
        cell = ws.cell(row=start, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
    for r_offset, row in enumerate(doc.rows, start=start + 1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_offset, column=c_idx, value=value)

    # Resumen debajo de la tabla.
    if doc.summary:
        base = start + 1 + len(doc.rows) + 1
        ws.cell(row=base, column=1, value="Resumen").font = gold
        for i, item in enumerate(doc.summary, start=1):
            ws.cell(row=base + i, column=1, value=item.label)
            ws.cell(row=base + i, column=2, value=item.value).font = Font(bold=True)

    # Ancho aproximado por contenido.
    for col_idx, name in enumerate(doc.columns, start=1):
        width = max(len(name), *(len(r[col_idx - 1]) for r in doc.rows if col_idx - 1 < len(r)), 8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 48)
    ws.cell(row=start, column=1).alignment = Alignment(horizontal="left")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── PDF ──────────────────────────────────────────────────────────────────────
_SUBS = {"Δ": "Var ", "«": '"', "»": '"', "≈": "~", "—": "-", "·": "-"}


def _pdf_safe(text: str) -> str:
    for bad, good in _SUBS.items():
        text = text.replace(bad, good)
    return text.encode("latin-1", "replace").decode("latin-1")


def to_pdf(doc: ExportDoc) -> bytes:
    from fpdf import FPDF

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    epw = pdf.epw  # ancho útil de página

    # Cabecera de marca.
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(122, 98, 40)
    pdf.cell(0, 9, _pdf_safe(doc.brand_name), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(40, 40, 40)
    pdf.cell(0, 7, _pdf_safe(doc.title), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(
        0,
        5,
        _pdf_safe(f"{doc.document_number}  ·  {doc.generated_at.isoformat(timespec='seconds')}"),
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.ln(2)

    ncols = max(len(doc.columns), 1)
    col_w = epw / ncols
    row_h = 7

    # Encabezado de tabla.
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(26, 26, 36)
    pdf.set_text_color(255, 255, 255)
    for name in doc.columns:
        pdf.cell(col_w, row_h, _pdf_safe(name)[:40], border=0, fill=True, align="L")
    pdf.ln(row_h)

    # Filas.
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(30, 30, 30)
    fill = False
    for row in doc.rows:
        if pdf.get_y() + row_h > pdf.h - 12:
            pdf.add_page()
        pdf.set_fill_color(245, 245, 247)
        for c in range(ncols):
            text = row[c] if c < len(row) else ""
            pdf.cell(col_w, row_h, _pdf_safe(text)[:40], border="B", fill=fill, align="L")
        pdf.ln(row_h)
        fill = not fill

    # Resumen.
    if doc.summary:
        pdf.ln(3)
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(122, 98, 40)
        pdf.cell(0, 6, "Resumen", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(40, 40, 40)
        for item in doc.summary:
            pdf.cell(0, 5, _pdf_safe(f"{item.label}: {item.value}"), new_x="LMARGIN", new_y="NEXT")

    return bytes(pdf.output())


# ── Respuesta HTTP ───────────────────────────────────────────────────────────
def render(doc: ExportDoc, fmt: ExportFormat) -> bytes:
    if fmt == "csv":
        return to_csv(doc)
    if fmt == "xlsx":
        return to_xlsx(doc)
    return to_pdf(doc)


def export_response(doc: ExportDoc, fmt: ExportFormat, *, filename_base: str) -> Response:
    content = render(doc, fmt)
    filename = f"{filename_base}.{fmt}"
    return Response(
        content=content,
        media_type=_MEDIA[fmt],
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
